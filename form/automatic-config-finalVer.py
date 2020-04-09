#!/usr/bin/python

print("Content-Type: text/plain\n")

from openpyxl import load_workbook
import cgi
import paramiko
import time

# Declaration des variables recuperees par le formulaire 
form = cgi.FieldStorage()
oldDeviceNum = form['oldDeviceNumber'].value
oldMACAddr = form['oldMACAddress'].value
oldNetSocket = form['oldNetworkSocket'].value
oldIPAddr = form['oldIPAddress'].value
oldSwitchPort = form['oldSwitchPort'].value
newNetSocket = form['newNetworkSocket'].value
newIPAddr = form['newIPAddress'].value
newSwitchPort = form['newSwitchPort'].value

# Declaration des variables de la documentation
# doc_path : Chemin + nom de la doc excel ou csv
doc_path = "/home/sharefolder/tab_exemple.xlsx" 
workbook = load_workbook(filename=doc_path)  
sheet = workbook.active

# Declaration des variables systemes impactees par le changement #

# VARIABLES DU SERVEUR DHCP #
# Adresse IP du serveur DHCP a renseigner pour la connnexion SSH
# Port de connexion SSH vers le serveur
# Nom d'utilisateur du serveur pour se connecter en SSH
# Mot de passe de l'utiisateur pour la connexion SSH
# Nom du serveur DHCP
# La plage DHCP dans lequel est fait le changement
DHCP_SERV_IP = "10.50.2.10" 
DHCP_CONNECT_PORT= 22
Windows_Serv_username = "Administrateur"
Windows_Serv_password = "Admin777"
DHCP_SERV_NAME = "DHCP-SERV" 
scopeID = "10.50.2.0"


### La fonction change_DHCP_Config modifie la reservation DHCP du serveur 
### en se connectant en SSH au serveur Windows et en supprimant la reservation DHCP actuelle 
### et la remplace par une nouvelle reservation avec les nouvelles infos.
### Le changement se fait par une commande powershell
def change_DHCP_Config(DHCP_IP,DHCP_SSH_PORT,Win_Serv_user,Win_Serv_pwd,DHCP_NAME,scopeIP,old_IP,old_MAC,Device_Num,new_IP):
	#run_powershell = "powershell -command "
	powershell_cmd_rm_DHCP = 'powershell -command Remove-DhcpServerv4Reservation -ScopeId ' + scopeIP + ' -ClientId ' + old_MAC + ' -ComputerName ' + DHCP_IP
	powershell_cmd_add_DHCP = 'powershell -command Add-DhcpServerv4Reservation -Name ' + Device_Num + ' -ScopeID ' + scopeIP + ' -IPAddress ' + new_IP + ' -ClientId ' + old_MAC + ' -Description ' + Device_Num
	#full_powershell_cmd = run_powershell + powershell_cmd_rm_DHCP + "\n" + run_powershell + powershell_cmd_add_DHCP + "\n" + "exit"; 
	ssh = paramiko.SSHClient()
	ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
	ssh.connect(hostname = DHCP_IP, port= DHCP_SSH_PORT, username = Win_Serv_user, password = Win_Serv_pwd)
	#full_powershell_cmd = run_powershell + powershell_cmd_rm_DHCP + "\n" + run_powershell + powershell_cmd_add_DHCP + "\n";
	stdin, stdout, stderr = ssh.exec_command(powershell_cmd_rm_DHCP)
	while not stdout.channel.exit_status_ready():
		if stdout.channel.recv_ready():
			stdoutLines = stdout.readlines()
	stdin, stdout, stderr = ssh.exec_command(powershell_cmd_add_DHCP)
	while not stdout.channel.exit_status_ready():
		if stdout.channel.recv_ready():
			stdoutLines = stdout.readlines()
	ssh.close()
	print(powershell_cmd_rm_DHCP)
	return;


def edit_documentation(newNetSocket,newSwitchPort,newIP):
	row_index = 1
	for row in sheet.iter_rows(): 
		if row[0].value == oldDeviceNum:
                	print(row[0].value + " " + str(row_index))
			row[2].value = newNetSocket
			row[3].value = newSwitchPort
			row[4].value = newIP
			workbook.save(filename=doc_path)
		row_index += 1
	return;

def read_doc():
        row_index = 1
	for row in sheet.iter_rows(values_only=True):
                if row[0] == oldDeviceNum:
			print(sheet["C" + str(row_index)].value)
		row_index +=1
	return;

change_DHCP_Config(DHCP_SERV_IP,DHCP_CONNECT_PORT,Windows_Serv_username,Windows_Serv_password,DHCP_SERV_NAME,scopeID,oldIPAddr,oldMACAddr,oldDeviceNum,newIPAddr)
edit_documentation(newNetSocket,newSwitchPort,newIPAddr)
#read_doc()


