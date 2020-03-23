#!/usr/bin/python

print("Content-Type: text/plain\n")

from openpyxl import load_workbook
import cgi

# Declaration des variables recuperees par le formulaire 
form = cgi.FieldStorage()
oldDeviceNum = form['oldDeviceNumber'].value
oldMACAddr = form['oldMACAddress'].value
oldNetSocket = form['oldNetworkSocket'].value
oldIPAddr = form['oldIPAddress'].value
oldSwitchPort = form['oldSwitchPort'].value
newNetSocket = form['newNetworkSocket'].value
newIPAddr = form['newIPAddress'].value
newSwitchPort = form['newSwitchPort'].value

# Declaration des variables de la documentation
# doc_path : Chemin + nom de la doc excel ou csv
doc_path = "/var/www/cgi-bin/tab_doc.xlsx" 
workbook = load_workbook(filename=doc_path)  
sheet = workbook.active

# Declaration des variables systemes impactees par le changement #

# VARIABLES DU SERVEUR DHCP #
# Adresse IP du serveur DHCP a renseigner pour la connnexion SSH
# Port de connexion SSH vers le serveur
# Nom d'utilisateur du serveur pour se connecter en SSH
# Mot de passe de l'utiisateur pour la connexion SSH
# Nom du serveur DHCP
# La plage DHCP dans lequel est fait le changement
DHCP_SERV_IP = "10.50.2.10" 
DHCP_CONNECT_PORT= 22
Windows_Serv_username = "Administrateur"
Windows_Serv_password = "Admin777"
DHCP_SERV_NAME = "DHCP-SERV" 
scopeID = "10.50.2.20"


### La fonction change_DHCP_Config modifie la reservation DHCP du serveur 
### en se connectant en SSH au serveur Windows et en supprimant la reservation DHCP actuelle 
### et la remplace par une nouvelle reservation avec les nouvelles infos.
### Le changement se fait par une commande powershell
def change_DHCP_Config(DHCP_IP,DHCP_SSH_PORT,Win_Serv_user,Win_Serv_pwd,DHCP_NAME,scopeIP,old_IP,old_MAC,Device_Num,new_IP):
	run_powershell = "powershell -command "
	powershell_cmd_rm_DHCP = 'Remove-DhcpServerv4Reservation -ComputerName ' + DHCP_NAME + ' -IPAddress ' + old_IP
	powershell_cmd_add_DHCP = 'Add-DhcpServerv4Reservation -Name ' + Device_Num + ' -ScopeID ' + scopeIP + ' -ClientId ' + old_MAC + ' -Description ' + Device_Num
	full_powershell_cmd = run_powershell + powershell_cmd_rm_DHCP + "\n" + run_powershell + powershell_cmd_add_DHCP + "\n" + "exit"; 
	#ssh = paramiko.SSHClient()
	#ssh.connect(DHCP_IP, DHCP_SSH_PORT, Win_Serv_user, Win_Serv_pwd)
	#stdin, stout, stderr = ssh.exec_command(full_powershell_cmd) 
	print(full_powershell_cmd)
	return;


def edit_documentation(newNetSocket,newSwitchPort,newIP):
	row_index = 1
	for row in sheet.iter_rows(values_only=True): 
		if row[0] == oldDeviceNum:
                	print(row[0] + " " + str(row_index))
			sheet["C" + str(row_index)] = newNetSocket
			sheet["D" + str(row_index)] = newSwitchPort
			sheet["E" + str(row_index)] = newIP
			workbook.save(doc_path)
		row_index += 1
	return;

def read_doc():
        row_index = 1
	for row in sheet.iter_rows(values_only=True):
                if row[0] == oldDeviceNum:
			print(sheet["C" + str(row_index)].value)
		row_index +=1
	return;

change_DHCP_Config(DHCP_SERV_IP,DHCP_CONNECT_PORT,Windows_Serv_username,Windows_Serv_password,DHCP_SERV_NAME,scopeID,oldIPAddr,oldMACAddr,oldDeviceNum,newIPAddr)
#edit_documentation(newNetSocket,newSwitchPort,newIPAddr)
read_doc()


#for row in sheet.iter_rows(values_only=True): 
#        if row[0] == oldDeviceNum:
#                print(row[0]) 

